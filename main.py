import os
from typing import List, Dict
import fitz
import openpyxl


def extract_blocks_from_pdf(file_path: str) -> List[Dict]:
    blocks = []
    current_block = {
        "authors": [],
        "affiliations": [],
        "session": "",
        "title": "",
        "abstract": "",
    }
    previous_block = None

    doc = fitz.open(file_path)

    for page_number in range(44 - 1, doc.page_count):
        page = doc.load_page(page_number)
        blocks_in_page = []
        text_blocks = page.get_text("dict")["blocks"]

        for block in text_blocks:
            for line in block["lines"]:
                for span in line["spans"]:
                    paragraph = (
                        span["font"] == "TimesNewRomanPS-BoldItal"
                        and span["size"] == 9.5
                    )
                    title = (
                        span["font"] == "TimesNewRomanPS-BoldMT"
                        and span["size"] == 9
                    )
                    authors = (
                        span["font"] == "TimesNewRomanPS-ItalicMT"
                        and span["size"] == 9
                    )
                    affiliations = (
                        span["font"] == "TimesNewRomanPS-ItalicMT"
                        and span["size"] == 8
                    )
                    abstract = span["size"] == 9.134002685546875

                    text_pattern = span["text"]

                    if paragraph:
                        current_block["session"] += text_pattern
                    elif title:
                        current_block["title"] += text_pattern
                    elif authors:
                        current_block["authors"].append(text_pattern)
                    elif affiliations:
                        current_block["affiliations"].append(text_pattern)
                    elif abstract:
                        current_block["abstract"] += text_pattern

            if block["type"] == 0:
                if (
                    current_block["session"]
                    or current_block["title"]
                    or current_block["authors"]
                    or current_block["affiliations"]
                    or current_block["abstract"]
                ):
                    if not current_block["session"]:
                        if previous_block is not None:
                            previous_block["title"] += current_block["title"]
                            previous_block["affiliations"].extend(
                                current_block["affiliations"]
                            )
                            previous_block["abstract"] += current_block[
                                "abstract"
                            ]
                            current_block = previous_block
                    blocks_in_page.append(current_block)
                    previous_block = current_block
                    current_block = {
                        "session": "",
                        "title": "",
                        "authors": [],
                        "affiliations": [],
                        "abstract": "",
                    }

        blocks.extend(blocks_in_page)

    doc.close()
    return blocks


def merge_blocks(blocks: List[Dict]) -> List[Dict]:
    merged_blocks = []
    current_block = None

    for block in blocks:
        if not current_block:
            current_block = block
        elif block["session"]:
            merged_blocks.append(current_block)
            current_block = block
        else:
            current_block["title"] += block["title"]
            current_block["affiliations"].extend(block["affiliations"])
            current_block["abstract"] += block["abstract"]

    if current_block:
        merged_blocks.append(current_block)

    return merged_blocks


def save_to_excel(blocks: List[Dict], file_path: str) -> None:
    if not os.path.exists(file_path):
        create_excel_file(file_path)

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    row = sheet.max_row + 1

    previous_session = ""
    previous_authors = []

    for block in blocks:
        session = block["session"]
        title = block["title"]
        authors = block["authors"]
        affiliations = block["affiliations"]
        abstract = block["abstract"]

        if session == previous_session and authors == previous_authors:
            continue

        for author in authors:
            if author.startswith(","):
                author = author[2:]
            author_exists = check_author_exists(sheet, session, author)

            if not author_exists:
                sheet.cell(row=row, column=1).value = author
                sheet.cell(row=row, column=2).value = ", ".join(affiliations)
                sheet.cell(row=row, column=3).value = session
                sheet.cell(row=row, column=4).value = title
                sheet.cell(row=row, column=5).value = abstract
                row += 1

        previous_session = session
        previous_authors = authors

    workbook.save(file_path)


def create_excel_file(file_path: str) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Author"
    sheet["B1"] = "Affiliations"
    sheet["C1"] = "Session"
    sheet["D1"] = "Title"
    sheet["E1"] = "Abstract"
    workbook.save(file_path)


def check_author_exists(sheet, session, author) -> bool:
    for row_num in range(2, sheet.max_row + 1):
        if (
            sheet.cell(row=row_num, column=3).value == session
            and sheet.cell(row=row_num, column=1).value == author
        ):
            return True
    return False


pdf_blocks = extract_blocks_from_pdf("beetroot_task.pdf")

merged_blocks = merge_blocks(pdf_blocks)

save_to_excel(merged_blocks, "result_.xlsx")
